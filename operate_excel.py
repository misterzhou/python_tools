#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author: mason
# Date: 2014-10-29
# Filename: operate_excel.py

'''
python读写Excel
Require: 需要先安装Python模块: xlrd xlwt xlutils
'''

__author__ = 'mason'

from xlrd import open_workbook
import xlwt
#from xlwt import easyxf
from xlutils.copy import copy
from openpyxl import Workbook

class Excel(object):

    #def __init__(self):

    def write(self, tpl, dest_file, tuple_args):
        '''
        Args:
          tuple_args: 从数据库获取的结果集(list)
          dest_file: 生成的excel文件名
          tpl: excel模板
        '''

        # 单元格样式
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = 'SimSun' #设置字体
        style.font = font
        tpl_book = open_workbook(tpl, formatting_info=True)
        target = copy(tpl_book)
        i = 0
        for arg in tuple_args:
            sheet = target.get_sheet(i) # a writable copy
            i += 1
            size = len(arg)
            if size == 0:
                continue
            cols = len(arg[0])
            for row in xrange(0,size):
                for col in xrange(0, cols):
                    sheet.write(row + 1, col, arg[row][col], style)
        target.save(dest_file)


    def write_with_openpyxl(self, dest_file, tuple_args):
        '''
        结果记录行数大于65536时，使用openpyxl模板生成Excel文件。
        xls文件的最大行数为65536，超过的话需用xlsx存储数据。
        注： openpyxl操作Excel时单元格起始下标为1。
        '''
        
        wb = Workbook() # 默认会创建一个sheet
        ws = wb.active
        isFirstSheet = False
        for arg in tuple_args:
            if isFirstSheet:
                ws = wb.create_sheet()
            isFirstSheet = True
            size = len(arg)
            if size == 0:
                continue
            cols = len(arg[0])
            header =['城市','板块','类型','社区名称','社区ID','用户数','设备数','新增用户数','新增设备数','日期']
            ws.append(header) #设置表头
            for r in xrange(2,size+1):
                for c in xrange(1, cols+1):
                    ws.cell(row=r, column=c).value = arg[r-2][c-1]
        wb.save(dest_file)

